import os
import re
import json
from openpyxl import load_workbook
import app.constants as constants

if __name__ == "__main__":
    path = os.path.join(constants.BUILD_SOURCES_DIR, constants.mitigations_mapping_dir)
    file = f"{path}/mappings.xlsx"

    workbook = load_workbook(file)
    
    uses_sheet = workbook["ATT&CK Mitigations v15"]
    mitre_sheet = workbook["MITRE Controls"]
    nist_sheet = workbook["NIST Controls"]
    ism_sheet = workbook["ISM Controls"]

    mitigations = {}
    uses = {}

    # mitre control uses
    overarching_mitigation = ""
    overarching_technique = ""

    for row_i, row in enumerate(uses_sheet.iter_rows()):
        if row_i == 1: continue

        technique = ""
        values = {}

        # 0: id, 1: sub-technique, 3: use
        for i, col in enumerate(row):
            col = str(col.value)

            if i == 0:
                mitigation = re.findall(r".*?\((M[0-9]{4})\)", col)
                technique = re.findall(r"(T[0-9]{4})", col)

                if len(technique) != 0: overarching_technique = technique[0]
                if len(mitigation) != 0: overarching_mitigation = mitigation[0]
            if i == 1 and row_i != 0:
                technique = overarching_technique

                if col != "None":
                    technique = f"{overarching_technique}.{col.split('.')[1]}"
            if i == 3 and row_i != 0:
                if not overarching_mitigation in uses:
                    uses[overarching_mitigation] = {}
                
                uses[overarching_mitigation][technique] = { "use": col }
    

    # mitre controls
    for i, rows in enumerate(mitre_sheet.iter_rows()):
        if i == 0: continue

        values = { "source": "MITRE" }

        # 0: id, 1: name, 2: description
        for i, row in enumerate(rows):
            if i == 1: values["name"] = row.value
            if i == 2: values["description"] = row.value
            if i == 0: 
                mitigations[row.value] = values
                
                if row.value in uses:
                    mitigations[row.value]["techniques"] = uses[row.value]

    # nist controls
    for i, rows in enumerate(nist_sheet.iter_rows()):
        if i == 0: continue

        values = { "source": "NIST" }

        # 0: id, 1: name, 2: description
        for i, row in enumerate(rows):

            # not a nist sub-technique
            if re.match(r"[A-Z]{2}-[0-9]{1,2}\([0-9]{1,2}\)", str(row.value)): continue
            
            if i == 1: values["name"] = row.value
            if i == 2: values["description"] = row.value
            if i == 0: mitigations[row.value] = values

    # ism controls
    for i, rows in enumerate(ism_sheet.iter_rows()):
        if i == 0: continue

        values = { "source": "ISM" }

        # 1: name, 3: id, 6: description
        for i, row in enumerate(rows):
            if i == 1: values["name"] = row.value
            if i == 6: values["description"] = row.value
            if i == 3: mitigations[row.value] = values

    with open(f"{path}/mappings-v15.1.json", "w") as outfile: 
        json.dump(mitigations, outfile)